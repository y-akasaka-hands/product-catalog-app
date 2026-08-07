import io
import math
import re
import zipfile
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import streamlit as st

# 画面全体のページ設定
st.set_page_config(page_title="共同販促売上集計プログラム", layout="wide", page_icon="📊")

# -----------------------------------------------------------------------------
# カスタムCSS: 濃いめのグリーン背景、高コントラスト＆くっきり枠線のカードUI
# -----------------------------------------------------------------------------
st.markdown("""
<style>
/* 全体の背景色をより濃くしっかりとしたグリーンに変更 */
.stApp {
    background-color: #d8ece3 !important;
}

/* ラジオボタンエリア（Step 1） */
div[data-testid="stRadio"] {
    background-color: #ffffff !important;
    padding: 20px 24px !important;
    border-radius: 12px !important;
    border: 2px solid #2e7d32 !important;
    box-shadow: 0 4px 10px rgba(0, 0, 0, 0.08) !important;
}

/* ファイルアップロード領域（Step 2 外枠＆内部ドロップエリア） */
div[data-testid="stFileUploader"] {
    background-color: #ffffff !important;
    padding: 16px !important;
    border-radius: 12px !important;
    border: 2px solid #2e7d32 !important;
    box-shadow: 0 4px 10px rgba(0, 0, 0, 0.08) !important;
}

div[data-testid="stFileUploaderDropzone"] {
    background-color: #f8fff9 !important;
    border: 2px dashed #2e7d32 !important;
    border-radius: 8px !important;
}

/* テキストインプット、ナンバーインプット＆セレクトボックス枠 (Step 3) */
div[data-testid="stTextInput"], div[data-testid="stNumberInput"], div[data-testid="stSelectbox"] {
    background-color: #ffffff !important;
    padding: 12px 16px !important;
    border-radius: 10px !important;
    border: 2px solid #2e7d32 !important;
    box-shadow: 0 4px 8px rgba(0, 0, 0, 0.06) !important;
    margin-bottom: 12px !important;
}

/* セクション見出しテキスト */
h5 {
    color: #1b5e20 !important;
    font-weight: 800 !important;
    font-size: 1.1rem !important;
    margin-bottom: 10px !important;
}

/* 立体的なボタン */
div.stDownloadButton > button, div.stButton > button {
    background: linear-gradient(180deg, #28a745 0%, #1e7e34 100%) !important;
    color: #ffffff !important;
    font-size: 18px !important;
    font-weight: 800 !important;
    padding: 14px 28px !important;
    border-radius: 8px !important;
    border: none !important;
    border-bottom: 4px solid #145222 !important;
    box-shadow: 0px 4px 8px rgba(0, 0, 0, 0.2) !important;
    transition: all 0.15s ease-in-out !important;
    width: 100% !important;
}

div.stDownloadButton > button:hover, div.stButton > button:hover {
    background: linear-gradient(180deg, #34ce57 0%, #218838 100%) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0px 6px 12px rgba(0, 0, 0, 0.25) !important;
    color: #ffffff !important;
}

div.stDownloadButton > button:active, div.stButton > button:active {
    transform: translateY(2px) !important;
    border-bottom: 2px solid #145222 !important;
    box-shadow: 0px 2px 4px rgba(0, 0, 0, 0.15) !important;
}
</style>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# ヘッダーエリア
# -----------------------------------------------------------------------------
st.title("📊 共同販促売上集計プログラム（店舗別・取引先別）")
st.caption("商品カタログ、共同販促パターン、POSデータの各種ファイルから取引先別・店舗別の集計表を自動生成します。")

# セッション状態の初期化
if "run_calc" not in st.session_state:
    st.session_state.run_calc = False
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0


# リセット処理関数
def reset_app():
    st.session_state.run_calc = False
    st.session_state.uploader_key += 1
    if f"sponsor_rate_{st.session_state.uploader_key-1}" in st.session_state:
        del st.session_state[f"sponsor_rate_{st.session_state.uploader_key-1}"]


# シート名・ファイル名修正関数
def clean_sheet_name(name):
    cleaned = re.sub(r"[\\/*?:\[\]]", "", str(name))
    return cleaned[:31] if cleaned else "Sheet"

def clean_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", str(name)).strip()


# 会員番号マスク関数（下4桁を残し前半を*で置換）
def mask_member_id(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if len(s) > 4:
        return "*" * (len(s) - 4) + s[-4:]
    return s


# JANコード13桁文字列化関数
def format_jan_13digits(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.replace(".0", "").isdigit():
        return str(int(float(s))).zfill(13)
    return s.zfill(13)


# -----------------------------------------------------------------------------
# セクション1: 集計パターンの選択
# -----------------------------------------------------------------------------
st.markdown("##### 📌 Step 1: 集計パターンの選択")
calc_mode = st.radio(
    "集計パターンを選択してください：",
    [
        "① 商品カタログ ＋ 店コード表（通常集計）",
        "② 共同販促パターン ＋ 商品カタログ ＋ 店コード表（共同販促集計）",
        "③ POSデータ ＋ 商品カタログ ＋ 店コード表（POS集計）"
    ],
    horizontal=True,
    label_visibility="collapsed",
    key=f"calc_mode_{st.session_state.uploader_key}"
)

# パターン③の場合のサブ選択
pos_sub_mode = "バンドル集計"
if "③" in calc_mode:
    st.markdown("###### 🔍 POS集計の種別を選択してください：")
    pos_sub_mode = st.radio(
        "POS集計種別",
        ["バンドル集計", "HC会員売上集計"],
        horizontal=True,
        label_visibility="collapsed",
        key=f"pos_sub_mode_{st.session_state.uploader_key}"
    )

st.markdown("<br>", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# セクション2: ファイル選択 ＆ 条件設定（左右2カラム配置）
# -----------------------------------------------------------------------------
col_file, col_opts = st.columns([3, 2], gap="large")

with col_file:
    st.markdown("##### 📂 Step 2: ファイルのアップロード")
    if "①" in calc_mode:
        file_help = "「商品カタログ」と「⑩店コード表」の2つのファイルをドラッグ＆ドロップしてください。"
    elif "②" in calc_mode:
        file_help = "「共同販促パターン」「商品カタログ」「⑩店コード表」の3つのファイルを同時にドラッグ＆ドロップしてください。"
    else:
        file_help = "「POSデータ」「商品カタログ」「⑩店コード表」の3つのファイルを同時にドラッグ＆ドロップしてください。"

    uploaded_files = st.file_uploader(
        file_help,
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key=f"file_uploader_{st.session_state.uploader_key}",
    )

with col_opts:
    st.markdown("##### ⚙️ Step 3: 企画情報・オプション設定")
    
    col_p1, col_p2 = st.columns(2)
    with col_p1:
        plan_category = st.text_input("企画区分", value="共同販促", key=f"plan_cat_{st.session_state.uploader_key}")
    with col_p2:
        plan_name = st.text_input("企画名", value="売上集計", key=f"plan_name_{st.session_state.uploader_key}")

    sponsor_rate_input = st.number_input(
        "協賛料率（％）",
        min_value=0.0,
        max_value=100.0,
        value=None,
        step=0.5,
        format="%.1f",
        key=f"sponsor_rate_{st.session_state.uploader_key}",
        help="例: 10 と入力すると 10% として計算されます（空欄可）",
    )

    if "②" in calc_mode:
        apportion_base = st.selectbox(
            "補填額（協賛額）の按分基準：",
            ["売上高", "売上原価", "付与ポイント"],
            key=f"apportion_base_{st.session_state.uploader_key}"
        )
    elif "③" in calc_mode:
        if pos_sub_mode == "バンドル集計":
            apportion_base = st.selectbox(
                "補填額（協賛額）の按分基準：",
                ["値引前売価", "値引後売価", "値引販売した商品の売上原価"],
                key=f"apportion_base_{st.session_state.uploader_key}"
            )
        else:
            apportion_base = st.selectbox(
                "補填額（協賛額）の按分基準：",
                ["HC会員売上", "HC会員売上原価", "付与ポイント"],
                key=f"apportion_base_{st.session_state.uploader_key}"
            )
    else:
        apportion_base = "売上高"

st.markdown("<br>", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 使い方ガイド（折りたたみ表示）
# -----------------------------------------------------------------------------
with st.expander("📖 詳しい使い方・仕様ガイドを見る"):
    st.markdown("""
    - **パターン①（通常集計）**: 「商品カタログ」「店コード表」の2ファイルが必要です。
    - **パターン②（共同販促集計）**: 「共同販促パターン」「商品カタログ」「店コード表」の3ファイルが必要です。
    - **パターン③（POSデータ集計）**: 「POSデータ」「商品カタログ」「店コード表」の3ファイルが必要です。
      - **バンドル集計**: `明細値引有無 = 0` の行を対象とし、割引率が協賛料率+3%以上の場合は売単価を再計算します。
      - **HC会員売上集計**: `HC番号 != _` の行を対象として会員売上を合算します。
    - **出力命名ルール**: `【企画区分】取引先名様_企画名_集計.xlsx` で取引先ごとに個別出力されます（複数取引先の場合はZIPファイル）。
    """)

# ----------------------------------------------------
# 処理実行ロジック
# ----------------------------------------------------
if uploaded_files:
    catalog_file = None
    store_code_file = None
    promo_file = None
    pos_file = None

    for f in uploaded_files:
        if "店コード" in f.name:
            store_code_file = f
        elif "共同販促" in f.name or "販促" in f.name:
            promo_file = f
        elif "pos" in f.name.lower():
            pos_file = f
        elif "商品カタログ" in f.name or "カタログ" in f.name:
            catalog_file = f

    ready_to_run = False
    if "①" in calc_mode:
        if not store_code_file: st.warning("⚠️ 「店コード表」が選択されていません。")
        if not catalog_file: st.warning("⚠️ 「商品カタログ」が選択されていません。")
        if store_code_file and catalog_file: ready_to_run = True
    elif "②" in calc_mode:
        if not promo_file: st.warning("⚠️ 「共同販促パターン」ファイルが選択されていません。")
        if not catalog_file: st.warning("⚠️ 「商品カタログ」ファイルが選択されていません。")
        if not store_code_file: st.warning("⚠️ 「店コード表」ファイルが選択されていません。")
        if promo_file and catalog_file and store_code_file: ready_to_run = True
    else:
        if not pos_file: st.warning("⚠️ 「POSデータ」ファイルが選択されていません。")
        if not catalog_file: st.warning("⚠️ 「商品カタログ」ファイルが選択されていません。")
        if not store_code_file: st.warning("⚠️ 「店コード表」ファイルが選択されていません。")
        if pos_file and catalog_file and store_code_file: ready_to_run = True

    if ready_to_run:
        if not st.session_state.run_calc:
            st.info("💡 ファイルの準備ができました。「集計を開始する」ボタンを押してください。")
            if st.button("🚀 集計を開始する", type="primary", use_container_width=True):
                st.session_state.run_calc = True
                st.rerun()

        if st.session_state.run_calc:
            try:
                # 店コード表読み込み
                store_map = {}
                fc_map = {}
                if store_code_file:
                    df_store_raw = pd.read_excel(store_code_file, header=None)
                    for idx, row in df_store_raw.iterrows():
                        if pd.notna(row[1]) and pd.notna(row[2]):
                            raw_code = str(row[1]).strip()
                            code = str(int(float(raw_code))).zfill(3) if raw_code.replace(".0", "").isdigit() else raw_code.zfill(3)
                            s_name = str(row[2]).strip()
                            store_map[s_name] = code
                            fc_val = str(row[4]).strip() if len(row) > 4 and pd.notna(row[4]) else ""
                            is_fc = "FC" if "FC" in fc_val.upper() else "直営"
                            fc_map[code] = is_fc
                            fc_map[s_name] = is_fc

                # 商品カタログ読み込み
                df_cat = pd.read_excel(catalog_file, header=None)
                header_row_idx = -1
                for i in range(min(30, len(df_cat))):
                    row_vals = [str(v).strip() for v in df_cat.iloc[i].values if pd.notna(v)]
                    if "売単価" in row_vals or "原単価" in row_vals or "商品名称" in row_vals or "商品ｺｰﾄﾞ" in row_vals:
                        header_row_idx = i
                        break

                if header_row_idx == -1:
                    st.error("❌ エラー: 商品カタログ内に見出し行が見つかりませんでした。")
                    st.stop()

                store_name_row_idx = max(0, header_row_idx - 1)
                row_store_names = df_cat.iloc[store_name_row_idx].values
                row_headers = df_cat.iloc[header_row_idx].values

                item_code_col, vendor_code_col, vendor_name_col, unit_price_col, cost_price_col, jan_col = -1, -1, -1, -1, -1, -1

                for idx, h_val in enumerate(row_headers):
                    if pd.isna(h_val): continue
                    h_str = str(h_val).strip()
                    if h_str == "売単価" and unit_price_col == -1: unit_price_col = idx
                    if h_str == "原単価" and cost_price_col == -1: cost_price_col = idx
                    if h_str in ["取引先名称", "取引先名"] and vendor_name_col == -1: vendor_name_col = idx
                    if h_str in ["商品ｺｰﾄﾞ", "商品コード", "JAN", "jan"] and jan_col == -1: jan_col = idx

                    top_h = str(row_store_names[idx]).strip() if pd.notna(row_store_names[idx]) else ""
                    if "取引先" in top_h and h_str in ["ｺｰﾄﾞ", "コード"] and vendor_code_col == -1: vendor_code_col = idx
                    if "品番" in top_h and h_str in ["ｺｰﾄﾞ", "コード"] and item_code_col == -1: item_code_col = idx

                if vendor_code_col == -1 and vendor_name_col != -1 and vendor_name_col > 0:
                    vendor_code_col = vendor_name_col - 1
                if item_code_col == -1:
                    for idx, h_val in enumerate(row_headers):
                        if str(h_val).strip() in ["品番", "品番ｺｰﾄﾞ", "品番コード"]:
                            item_code_col = idx; break

                df_cat_data = df_cat.iloc[header_row_idx + 1:].copy()
                rate_val = (sponsor_rate_input / 100.0) if sponsor_rate_input is not None else None

                # 商品カタログマッピング作成
                jan_map = {}
                for _, row in df_cat_data.iterrows():
                    j_val = str(row[jan_col]).strip() if pd.notna(row[jan_col]) else ""
                    if j_val:
                        if j_val.endswith(".0"): j_val = j_val[:-2]
                        v_code = str(int(float(row[vendor_code_col]))).zfill(7) if pd.notna(row[vendor_code_col]) and str(row[vendor_code_col]).replace('.0','').isdigit() else str(row[vendor_code_col]).strip().zfill(7) if pd.notna(row[vendor_code_col]) else "0000000"
                        v_name = str(row[vendor_name_col]).strip() if pd.notna(row[vendor_name_col]) else ""
                        i_code = str(int(float(row[item_code_col]))) if pd.notna(row[item_code_col]) and str(row[item_code_col]).replace('.0','').isdigit() else str(row[item_code_col]).strip() if pd.notna(row[item_code_col]) else ""
                        u_price = float(row[unit_price_col]) if unit_price_col != -1 and pd.notna(row[unit_price_col]) else 0.0
                        c_price = float(row[cost_price_col]) if cost_price_col != -1 and pd.notna(row[cost_price_col]) else 0.0
                        jan_map[j_val] = {
                            "vendor_code": v_code,
                            "vendor_name": v_name,
                            "item_code": i_code,
                            "unit_price": u_price,
                            "cost_price": c_price
                        }

                df_raw_export = None
                raw_export_type = ""

                # ====================================================
                # パターン①：通常集計
                # ====================================================
                if "①" in calc_mode:
                    sales_col_header = "売上高\n(売単価×数量)"
                    sales_preview_header = "売上高（売単価×数量）"
                    unit_prices = pd.to_numeric(df_cat_data[unit_price_col], errors="coerce").fillna(0)

                    store_columns = []
                    for col_idx in range(len(row_store_names)):
                        top_val = str(row_store_names[col_idx]).strip() if pd.notna(row_store_names[col_idx]) else ""
                        if top_val != "" and top_val not in ["品番", "枝番", "取引先", "全店"]:
                            for offset in range(3):
                                check_idx = col_idx + offset
                                if check_idx < len(row_headers):
                                    sub_val = str(row_headers[check_idx]).strip() if pd.notna(row_headers[check_idx]) else ""
                                    if sub_val == "売上数":
                                        store_columns.append((top_val, check_idx))
                                        break

                    df_cat_data["vendor_code"] = df_cat_data[vendor_code_col].apply(
                        lambda x: str(int(float(x))).zfill(7) if pd.notna(x) and str(x).replace(".0", "").isdigit()
                        else str(x).strip().zfill(7) if pd.notna(x) else "0000000"
                    )
                    df_cat_data["vendor_name"] = df_cat_data[vendor_name_col].astype(str).str.strip()
                    df_cat_data["item_code"] = df_cat_data[item_code_col].apply(
                        lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace(".0", "").isdigit()
                        else str(x).strip() if pd.notna(x) else ""
                    )

                    records = []
                    for store_name, col_idx in store_columns:
                        qty_series = pd.to_numeric(df_cat_data[col_idx], errors="coerce").fillna(0)
                        sales_series = qty_series * unit_prices
                        store_code = store_map.get(store_name, "999")
                        temp_df = pd.DataFrame({
                            "vendor_code": df_cat_data["vendor_code"],
                            "vendor_name": df_cat_data["vendor_name"],
                            "store_code": store_code,
                            "store_name": store_name,
                            "item_code": df_cat_data["item_code"],
                            "sales": sales_series,
                            "apportion_val": sales_series
                        })
                        records.append(temp_df[temp_df["sales"] > 0])

                    df_detail_all = pd.concat(records, ignore_index=True) if records else pd.DataFrame()
                    df_detail_grouped = df_detail_all.groupby(
                        ["vendor_code", "vendor_name", "store_code", "store_name", "item_code"], as_index=False
                    ).agg({"sales": "sum", "apportion_val": "sum"})

                # ====================================================
                # パターン②：共同販促集計
                # ====================================================
                elif "②" in calc_mode:
                    sales_col_header = "HC会員売上\n(売単価×数量)"
                    sales_preview_header = "HC会員売上（売単価×数量）"
                    raw_export_type = "HC会員売上"

                    df_prom_raw = pd.read_excel(promo_file)
                    df_prom = df_prom_raw.copy()

                    master_price_col = next((c for c in df_prom.columns if "マスター単価" in str(c) or "単価" in str(c)), None)
                    qty_col = next((c for c in df_prom.columns if "数量" in str(c)), None)
                    store_code_col_prom = next((c for c in df_prom.columns if "店舗コード" in str(c) or "店コード" in str(c)), None)
                    store_name_col_prom = next((c for c in df_prom.columns if "店舗名" in str(c) or "店名" in str(c)), None)
                    jan_col_prom = next((c for c in df_prom.columns if "jan" in str(c).lower() or "商品コード" in str(c)), None)
                    item_code_col_prom = next((c for c in df_prom.columns if "品番" in str(c)), None)
                    point_col_prom = next((c for c in df_prom.columns if "付与ポイント" in str(c) or "ポイント" in str(c)), None)
                    member_col_prom = next((c for c in df_prom.columns if "会員" in str(c)), None)

                    df_prom["store_code_clean"] = df_prom[store_code_col_prom].apply(
                        lambda x: str(int(float(x))).zfill(3) if pd.notna(x) and str(x).replace(".0","").isdigit() else str(x).strip().zfill(3) if pd.notna(x) else "999"
                    )
                    df_prom_calc = df_prom[df_prom["store_code_clean"] != "052"].copy()
                    df_prom_calc["jan_str"] = df_prom_calc[jan_col_prom].apply(
                        lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace(".0","").isdigit() else str(x).strip() if pd.notna(x) else ""
                    )

                    df_prom_calc["sales"] = pd.to_numeric(df_prom_calc[master_price_col], errors="coerce").fillna(0) * pd.to_numeric(df_prom_calc[qty_col], errors="coerce").fillna(0)
                    df_prom_calc["catalog_unit_price"] = df_prom_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("unit_price", 0.0))
                    df_prom_calc["gross_sales"] = df_prom_calc["catalog_unit_price"] * pd.to_numeric(df_prom_calc[qty_col], errors="coerce").fillna(0)
                    df_prom_calc["cost_price"] = df_prom_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("cost_price", 0.0))
                    df_prom_calc["cost_total"] = df_prom_calc["cost_price"] * pd.to_numeric(df_prom_calc[qty_col], errors="coerce").fillna(0)

                    if point_col_prom and point_col_prom in df_prom_calc.columns:
                        df_prom_calc["points_clean"] = df_prom_calc[point_col_prom].apply(
                            lambda x: math.floor(float(x)) if pd.notna(x) and not math.isnan(float(x)) else 0
                        )
                    else:
                        df_prom_calc["points_clean"] = 0

                    df_prom_calc["store_name"] = df_prom_calc[store_name_col_prom].astype(str).str.strip()
                    df_prom_calc["vendor_code"] = df_prom_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("vendor_code", "0000000"))
                    df_prom_calc["vendor_name"] = df_prom_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("vendor_name", "不明"))
                    df_prom_calc["item_code"] = df_prom_calc.apply(
                        lambda r: jan_map.get(r["jan_str"], {}).get("item_code", str(r[item_code_col_prom]) if item_code_col_prom and pd.notna(r[item_code_col_prom]) else ""), axis=1
                    )

                    df_detail_grouped = df_prom_calc.groupby(
                        ["vendor_code", "vendor_name", "store_code_clean", "store_name", "item_code"], as_index=False
                    ).agg({
                        "sales": "sum",
                        "gross_sales": "sum",
                        "cost_total": "sum",
                        "points_clean": "sum"
                    }).rename(columns={"store_code_clean": "store_code"})

                    if apportion_base == "売上原価": df_detail_grouped["apportion_val"] = df_detail_grouped["cost_total"]
                    elif apportion_base == "付与ポイント": df_detail_grouped["apportion_val"] = df_detail_grouped["points_clean"]
                    else: df_detail_grouped["apportion_val"] = df_detail_grouped["sales"]

                    # 原本エクスポート準備
                    df_promo_export = df_prom_raw.copy()
                    if member_col_prom and member_col_prom in df_promo_export.columns:
                        df_promo_export[member_col_prom] = df_promo_export[member_col_prom].apply(mask_member_id)
                    if store_code_col_prom and store_code_col_prom in df_promo_export.columns:
                        df_promo_export[store_code_col_prom] = df_prom["store_code_clean"]
                    if jan_col_prom and jan_col_prom in df_promo_export.columns:
                        df_promo_export[jan_col_prom] = df_promo_export[jan_col_prom].apply(format_jan_13digits)

                    cols_to_drop = [c for c in df_promo_export.columns if "ポイント率" in str(c) or "ポイント計算基準額" in str(c)]
                    if cols_to_drop: df_promo_export = df_promo_export.drop(columns=cols_to_drop)
                    df_raw_export = df_promo_export

                # ====================================================
                # パターン③：POSデータ集計
                # ====================================================
                else:
                    raw_export_type = "POSデータ"
                    df_pos_raw = pd.read_excel(pos_file, header=None)
                    pos_headers_row = df_pos_raw.iloc[5].values
                    df_pos_data = df_pos_raw.iloc[6:].copy().reset_index(drop=True)
                    df_pos_data.columns = [str(c).strip() for c in pos_headers_row]

                    df_pos_data["store_code_clean"] = df_pos_data["店舗コード"].apply(
                        lambda x: str(int(float(x))).zfill(3) if pd.notna(x) and str(x).replace(".0","").isdigit() else str(x).strip().zfill(3) if pd.notna(x) else "999"
                    )
                    df_pos_calc = df_pos_data[df_pos_data["store_code_clean"] != "052"].copy()
                    df_pos_calc["jan_str"] = df_pos_calc["商品コード"].apply(
                        lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace(".0","").isdigit() else str(x).strip() if pd.notna(x) else ""
                    )

                    df_pos_calc["qty"] = pd.to_numeric(df_pos_calc["数量"], errors="coerce").fillna(0)
                    df_pos_calc["master_price"] = pd.to_numeric(df_pos_calc["マスタ売単価"], errors="coerce").fillna(0)
                    df_pos_calc["unit_price"] = pd.to_numeric(df_pos_calc["売単価"], errors="coerce").fillna(0)

                    df_pos_calc["catalog_unit_price"] = df_pos_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("unit_price", 0.0))
                    df_pos_calc["cost_price"] = df_pos_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("cost_price", 0.0))
                    df_pos_calc["gross_sales"] = df_pos_calc["catalog_unit_price"] * df_pos_calc["qty"]
                    df_pos_calc["cost_total"] = df_pos_calc["cost_price"] * df_pos_calc["qty"]

                    df_pos_calc["vendor_code"] = df_pos_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("vendor_code", "0000000"))
                    df_pos_calc["vendor_name"] = df_pos_calc["jan_str"].apply(lambda j: jan_map.get(j, {}).get("vendor_name", "不明"))
                    df_pos_calc["item_code"] = df_pos_calc.apply(
                        lambda r: jan_map.get(r["jan_str"], {}).get("item_code", str(r["品番コード"]) if "品番コード" in r and pd.notna(r["品番コード"]) else ""), axis=1
                    )
                    df_pos_calc["store_name"] = df_pos_calc["store_code_clean"].apply(lambda c: next((k for k, v in store_map.items() if v == c), f"店舗{c}"))

                    if pos_sub_mode == "バンドル集計":
                        sales_col_header = "値引後売価\n(売単価×数量)"
                        sales_preview_header = "値引後売価（売単価×数量）"

                        df_pos_sub = df_pos_calc[df_pos_calc["明細値引有無"].astype(str).str.strip() == "0"].copy()
                        df_pos_sub["discount_rate"] = df_pos_sub.apply(
                            lambda r: (r["master_price"] - r["unit_price"]) / r["master_price"] if r["master_price"] > 0 else 0.0, axis=1
                        )
                        def calc_bundle_price(r):
                            if rate_val is not None and r["master_price"] > 0:
                                if r["discount_rate"] >= (rate_val + 0.03):
                                    return r["master_price"] * (1.0 - rate_val)
                            return r["unit_price"]

                        df_pos_sub["final_unit_price"] = df_pos_sub.apply(calc_bundle_price, axis=1)
                        df_pos_sub["price_before_discount"] = df_pos_sub["master_price"] * df_pos_sub["qty"]
                        df_pos_sub["price_after_discount"] = df_pos_sub["final_unit_price"] * df_pos_sub["qty"]
                        df_pos_sub["sales"] = df_pos_sub["price_after_discount"]

                        df_detail_grouped = df_pos_sub.groupby(
                            ["vendor_code", "vendor_name", "store_code_clean", "store_name", "item_code"], as_index=False
                        ).agg({
                            "price_before_discount": "sum",
                            "price_after_discount": "sum",
                            "sales": "sum",
                            "cost_total": "sum",
                            "gross_sales": "sum"
                        }).rename(columns={"store_code_clean": "store_code"})

                        if apportion_base == "値引前売価": df_detail_grouped["apportion_val"] = df_detail_grouped["price_before_discount"]
                        elif apportion_base == "値引販売した商品の売上原価": df_detail_grouped["apportion_val"] = df_detail_grouped["cost_total"]
                        else: df_detail_grouped["apportion_val"] = df_detail_grouped["price_after_discount"]

                    else:
                        sales_col_header = "HC会員売上\n(売単価×数量)"
                        sales_preview_header = "HC会員売上（売単価×数量）"

                        df_pos_sub = df_pos_calc[df_pos_calc["HC番号"].astype(str).str.strip() != "_"].copy()
                        df_pos_sub["sales"] = df_pos_sub["unit_price"] * df_pos_sub["qty"]
                        df_pos_sub["points_clean"] = pd.to_numeric(df_pos_sub["HC付与ポイント数"], errors="coerce").fillna(0).apply(math.floor)

                        df_detail_grouped = df_pos_sub.groupby(
                            ["vendor_code", "vendor_name", "store_code_clean", "store_name", "item_code"], as_index=False
                        ).agg({
                            "sales": "sum",
                            "cost_total": "sum",
                            "gross_sales": "sum",
                            "points_clean": "sum"
                        }).rename(columns={"store_code_clean": "store_code"})

                        if apportion_base == "HC会員売上原価": df_detail_grouped["apportion_val"] = df_detail_grouped["cost_total"]
                        elif apportion_base == "付与ポイント": df_detail_grouped["apportion_val"] = df_detail_grouped["points_clean"]
                        else: df_detail_grouped["apportion_val"] = df_detail_grouped["sales"]

                    # POS原本フィルタリングエクスポートデータ作成（重複ヘッダーを防止）
                    keep_cols = [1, 23, 24, 48, 50, 51, 52, 53, 55, 56, 57, 58]
                    df_pos_export_raw = df_pos_raw.iloc[6:].copy().reset_index(drop=True)
                    df_pos_export = df_pos_export_raw.iloc[:, keep_cols].copy()
                    df_pos_export.columns = [pos_headers_row[i] for i in keep_cols]

                    df_pos_export["HC番号"] = df_pos_export["HC番号"].apply(lambda x: mask_member_id(x) if str(x).strip() != "_" else "_")
                    df_pos_export["店舗コード"] = df_pos_export["店舗コード"].apply(
                        lambda x: str(int(float(x))).zfill(3) if pd.notna(x) and str(x).replace(".0","").isdigit() else str(x).strip().zfill(3) if pd.notna(x) else "999"
                    )
                    df_raw_export = df_pos_export

                # ====================================================
                # 取引先別ファイルの生成処理 (Excel / ZIP)
                # ====================================================
                st.success(f"✅ 集計が完了しました！（按分基準: {apportion_base}）")
                tab1, tab2 = st.tabs(["🏬 全店サマリー確認", "🏢 取引先別・明細プレビュー"])

                # プレビュー表
                with tab1:
                    st.markdown("##### 🏬 店舗別サマリープレビュー")
                    df_sum_prev = df_detail_grouped.groupby(["store_code", "store_name"], as_index=False)["sales"].sum()
                    st.dataframe(df_sum_prev.style.format({"sales": "{:,.0f}"}), use_container_width=True)

                with tab2:
                    st.markdown("##### 🏢 明細プレビュー")
                    st.dataframe(df_detail_grouped.style.format({"sales": "{:,.0f}"}), use_container_width=True)

                st.divider()

                # ZIP / 個別Excel出力ファイルの生成
                unique_vendors = df_detail_grouped[["vendor_code", "vendor_name"]].drop_duplicates()
                generated_files = [] # list of (filename, bytes_data)

                for _, v_row in unique_vendors.iterrows():
                    v_code, v_name = v_row["vendor_code"], v_row["vendor_name"]
                    v_clean_name = clean_filename(v_name)

                    # 出力ファイル名フォーマット: 【企画区分】取引先名様_企画名_集計.xlsx
                    out_excel_filename = f"【{clean_filename(plan_category)}】{v_clean_name}様_{clean_filename(plan_name)}_集計.xlsx"

                    df_v = df_detail_grouped[
                        (df_detail_grouped["vendor_code"] == v_code) & (df_detail_grouped["vendor_name"] == v_name)
                    ].copy()

                    # 店舗別サマリー
                    v_store_summary = df_v.groupby(["store_code", "store_name"], as_index=False).agg({
                        "sales": "sum",
                        "apportion_val": "sum"
                    })
                    if "gross_sales" in df_v.columns:
                        v_store_summary["gross_sales"] = df_v.groupby(["store_code", "store_name"])["gross_sales"].sum().values

                    v_total_sales = v_store_summary["sales"].sum()
                    v_total_apportion = v_store_summary["apportion_val"].sum()

                    v_store_sponsors = [round(amt * rate_val) for amt in v_store_summary["apportion_val"]] if rate_val is not None else []
                    if rate_val is not None:
                        v_target_sponsor = round(v_total_apportion * rate_val)
                        v_diff = v_target_sponsor - sum(v_store_sponsors)
                        if v_diff != 0 and len(v_store_sponsors) > 0:
                            max_idx = v_store_summary["apportion_val"].idxmax()
                            v_store_sponsors[max_idx] += v_diff

                    v_store_summary["構成比"] = v_store_summary["apportion_val"] / v_total_apportion if v_total_apportion > 0 else 0.0
                    v_store_summary["協賛額"] = v_store_sponsors if rate_val is not None else [None] * len(v_store_summary)
                    v_store_summary["協賛料率"] = [None] * len(v_store_summary)

                    # 全店行
                    v_tot_row = {
                        "店コード": "000",
                        "店舗名": "全店",
                        sales_col_header: v_total_sales,
                        "構成比": 1.0,
                        "協賛額": sum(v_store_sponsors) if rate_val is not None else None,
                        "協賛料率": rate_val if rate_val is not None else None
                    }
                    if "gross_sales" in v_store_summary.columns:
                        v_tot_row["総売上\n(売単価×数量)"] = v_store_summary["gross_sales"].sum()

                    df_v_store_res = pd.DataFrame([v_tot_row])

                    df_v_store_rows = pd.DataFrame({
                        "店コード": v_store_summary["store_code"],
                        "店舗名": v_store_summary["store_name"],
                        sales_col_header: v_store_summary["sales"],
                        "構成比": v_store_summary["構成比"],
                        "協賛額": v_store_summary["協賛額"],
                        "協賛料率": v_store_summary["協賛料率"]
                    })
                    if "gross_sales" in v_store_summary.columns:
                        df_v_store_rows["総売上\n(売単価×数量)"] = v_store_summary["gross_sales"]

                    df_v_store_final = pd.concat([df_v_store_res, df_v_store_rows], ignore_index=True)

                    # 取引先直営/FC集計
                    v_store_summary["fc_type"] = v_store_summary["store_code"].apply(lambda c: fc_map.get(c, "直営"))
                    v_chokuei_sales = v_store_summary[v_store_summary["fc_type"] == "直営"]["sales"].sum()
                    v_chokuei_sponsor = sum([v_store_sponsors[i] for i, r in v_store_summary.iterrows() if r["fc_type"] == "直営"]) if rate_val is not None else 0
                    v_fc_sales = v_store_summary[v_store_summary["fc_type"] == "FC"]["sales"].sum()
                    v_fc_sponsor = sum([v_store_sponsors[i] for i, r in v_store_summary.iterrows() if r["fc_type"] == "FC"]) if rate_val is not None else 0

                    # 取引先商品コードリスト作成
                    v_jan_list = [j for j, info in jan_map.items() if info.get("vendor_code") == v_code or info.get("vendor_name") == v_name]

                    excel_out = io.BytesIO()
                    with pd.ExcelWriter(excel_out, engine="openpyxl") as writer:
                        # シート1: 店舗別集計
                        df_v_store_final.to_excel(writer, index=False, sheet_name="店舗別集計")

                        # シート2: 取引先別明細
                        df_v_export_detail = df_v[["vendor_code", "vendor_name", "store_code", "store_name", "item_code", "sales"]].copy()
                        df_v_export_detail.columns = ["取引先コード", "取引先名", "店コード", "店舗名", "品番", sales_preview_header]
                        df_v_export_detail.to_excel(writer, index=False, sheet_name="取引先別明細")

                        # シート3: 原本データ（該当取引先の商品のみフィルタリング・重複ヘッダーなし）
                        if df_raw_export is not None:
                            if raw_export_type == "POSデータ":
                                df_pos_filt = df_raw_export[df_raw_export["商品コード"].astype(str).str.strip().isin(v_jan_list)]
                                df_pos_filt.to_excel(writer, index=False, sheet_name="POSデータ")
                            elif raw_export_type == "HC会員売上":
                                jan_col_name = next((c for c in df_raw_export.columns if "jan" in str(c).lower() or "商品コード" in str(c)), None)
                                if jan_col_name:
                                    df_hc_filt = df_raw_export[df_raw_export[jan_col_name].astype(str).str.strip().isin(v_jan_list)]
                                else:
                                    df_hc_filt = df_raw_export
                                df_hc_filt.to_excel(writer, index=False, sheet_name="HC会員売上")

                        # シート4: 商品カタログ原本（該当取引先のみフィルタリング）
                        df_cat_v = df_cat_data[
                            (df_cat_data[vendor_code_col].astype(str).str.strip() == v_code) |
                            (df_cat_data[vendor_name_col].astype(str).str.strip() == v_name)
                        ]
                        df_cat_export_v = pd.concat([df_cat.iloc[:header_row_idx + 1], df_cat_v], ignore_index=True)
                        df_cat_export_v.to_excel(writer, index=False, header=False, sheet_name="商品カタログ")

                        # 装飾適用
                        ws = writer.sheets["店舗別集計"]
                        FONT_NAME = "メイリオ"
                        header_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        body_font = Font(name=FONT_NAME, size=10)
                        total_font = Font(name=FONT_NAME, size=10, bold=True)
                        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        thin_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
                        total_bottom_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="double", color="000000"))

                        # 直営/FC小表追加
                        start_col = 9 if "総売上\n(売単価×数量)" in df_v_store_final.columns else 8
                        headers_fc = ["区分", "売上高", "協賛額"]
                        for idx_fc, h_text in enumerate(headers_fc):
                            c = ws.cell(row=1, column=start_col + idx_fc, value=h_text)
                            c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center", vertical="center")

                        fc_rows = [
                            ("直営店", v_chokuei_sales, v_chokuei_sponsor if rate_val is not None else None),
                            ("FC店", v_fc_sales, v_fc_sponsor if rate_val is not None else None),
                            ("合計", v_chokuei_sales + v_fc_sales, (v_chokuei_sponsor + v_fc_sponsor) if rate_val is not None else None),
                        ]
                        for r_idx, (cat_label, val_sales, val_spons) in enumerate(fc_rows, start=2):
                            is_tot = (cat_label == "合計")
                            curr_f = total_font if is_tot else body_font
                            curr_b = total_bottom_border if is_tot else thin_border

                            c1 = ws.cell(row=r_idx, column=start_col, value=cat_label)
                            c1.font = curr_f; c1.border = curr_b; c1.alignment = Alignment(horizontal="center")

                            c2 = ws.cell(row=r_idx, column=start_col + 1, value=val_sales)
                            c2.font = curr_f; c2.border = curr_b; c2.alignment = Alignment(horizontal="right"); c2.number_format = "#,##0"

                            c3 = ws.cell(row=r_idx, column=start_col + 2, value=val_spons)
                            c3.font = curr_f; c3.border = curr_b; c3.alignment = Alignment(horizontal="right")
                            if rate_val is not None: c3.number_format = "#,##0"
                            else: c3.value = None

                        # ヘッダー・セル装飾
                        for col_num in range(1, ws.max_column + 1):
                            if col_num >= start_col: continue
                            cell = ws.cell(row=1, column=col_num)
                            cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        for row_num in range(2, ws.max_row + 1):
                            is_total_row = (row_num == 2)
                            curr_f = total_font if is_total_row else body_font
                            curr_b = total_bottom_border if is_total_row else thin_border
                            for c in range(1, start_col):
                                cell = ws.cell(row=row_num, column=c)
                                cell.font = curr_f; cell.border = curr_b
                                if c == 1: cell.alignment = Alignment(horizontal="center")
                                elif c == 2: cell.alignment = Alignment(horizontal="left")
                                elif c in [3, 5, 7]: cell.alignment = Alignment(horizontal="right"); cell.number_format = "#,##0"
                                elif c == 4: cell.alignment = Alignment(horizontal="right"); cell.number_format = "0.0%"
                                elif c == 6:
                                    cell.alignment = Alignment(horizontal="right")
                                    if is_total_row and rate_val is not None: cell.number_format = "0%"
                                    else: cell.value = None

                        ws.cell(row=2, column=5).fill = yellow_fill

                        for col in ws.columns:
                            col_letter = openpyxl.utils.get_column_letter(col[0].column)
                            max_len = 0
                            for cell in col:
                                val_str = str(cell.value or "")
                                for l in val_str.split("\n"):
                                    length = sum(2 if ord(ch) > 256 else 1 for ch in l)
                                    if length > max_len: max_len = length
                            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

                    generated_files.append((out_excel_filename, excel_out.getvalue()))

                # ダウンロードエリアの構築
                col_dl, col_rst = st.columns([1, 1])

                if len(generated_files) == 1:
                    f_name, f_data = generated_files[0]
                    with col_dl:
                        st.download_button(
                            label="📥 集計結果（Excel）をダウンロード",
                            data=f_data,
                            file_name=f_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                else:
                    zip_buffer = io.BytesIO()
                    zip_filename = f"【{clean_filename(plan_category)}】{clean_filename(plan_name)}_取引先別集計.zip"

                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for f_name, f_data in generated_files:
                            zip_file.writestr(f_name, f_data)

                    with col_dl:
                        st.download_button(
                            label=f"📦 全取引先分（ZIPファイル/計{len(generated_files)}社）をダウンロード",
                            data=zip_buffer.getvalue(),
                            file_name=zip_filename,
                            mime="application/zip",
                            use_container_width=True,
                        )

                with col_rst:
                    st.button("🔄 最初の画面に戻る（全リセット）", on_click=reset_app, use_container_width=True)

            except Exception as e:
                st.error(f"❌ 予期せぬエラーが発生しました: {e}")
